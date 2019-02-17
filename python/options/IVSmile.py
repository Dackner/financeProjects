import numpy as np
from scipy.stats import norm
import random as rnd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from BlackScholesMertonFuncs import BlackScholesMerton

###### read data
wb = load_workbook(filename = 'plainVanillaData.xlsx')
sheet = wb['snap']
S = sheet.cell(row=1,column=2).value
K = [sheet.cell(row=7,column=i).value for i in range(2,6)]
T = sheet.cell(row=11,column=2).value
r = 0.01
q = 0


print(S)
print(K)
print(T)



S = 100
K = 100
T = 3/12
r = 0.01
q = 0
sigma = 0.2


C = BlackScholesMerton(S, K, T, r, q, sigma)

# print(C)